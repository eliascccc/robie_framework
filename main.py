import time, threading, traceback, os, tempfile, sys, platform, subprocess, signal, atexit, sqlite3, datetime, shutil, re, json
import tkinter as tk
from openpyxl import Workbook, load_workbook #type: ignore
from typing import Never, Literal, Any, get_args, TypeAlias
from pathlib import Path
from email.parser import BytesParser
from email.utils import parseaddr
from dataclasses import dataclass, asdict
from email import policy
#ctrl + k + 2 minimize. ctrl+k ctrl+j max


# =========================
# CUSTOMIZE HERE
# =========================
# Replace/adapt these parts for your environment:
# - ExampleMailBackend
# - ExampleErpBackend
# - ExampleJob*Handler classes
# - NetworkService.NETWORK_HEALTHCHECK_PATH
# - FriendsRepository file format if needed
# - RPA tool implementation that reads/writes handover.json (not in this code)


# ============================================================
# DATA MODELS
# ============================================================

IpcState: TypeAlias = Literal["idle", "job_queued", "job_running", "job_verifying", "safestop"]
JobType: TypeAlias = Literal["ping", "job1", "job2", "job3", "job4"]
JobSourceType: TypeAlias = Literal["personal_inbox", "shared_inbox", "erp_query"]
JobStatus: TypeAlias = Literal["REJECTED", "QUEUED", "RUNNING", "VERIFYING", "FAIL", "DONE"]
JobAction: TypeAlias = Literal["DELETE_ONLY", "REPLY_AND_DELETE", "QUEUE_RPA_JOB", "SKIP", "MOVE_BACK_TO_INBOX", "CRASH"]
UIStatusText: TypeAlias = Literal["online", "safestop", "working", "no network" , "ooo"]


@dataclass
class JobCandidate:
    source_ref: str
    job_source_type: JobSourceType
    source_data: dict[str, Any]

    sender_email: str | None = None # for email only
    subject: str | None = None  # for email only
    body: str | None = None  # for email only


@dataclass
class JobDecision:
    action: JobAction
    job_type: JobType | None = None
    job_status: JobStatus | None = None
    error_code: str | None = None
    error_message: str | None = None
    rpa_payload: dict[str, Any] | None = None
    ui_log_message: str | None = None
    system_log_message: str | None = None
    send_lifesign_notice: bool = False
    start_recording: bool = False



@dataclass
class ActiveJob:
    ''' goes into handover.json '''
    
    # common fields
    ipc_state: IpcState

    source_ref: str | None = None  # identifier, eg. "ERP_ORDER:12345" or "mail1234.eml"

    job_type: JobType | None = None
    job_source_type: JobSourceType | None = None
    job_id: int | None = None

    sender_email: str | None = None # for email
    subject: str | None = None      # for email
    body: str | None = None         # for email eg. "Hi, change the order 12345 to 44 pcs"

    # parsed from source 
    source_data: dict[str, Any] | None = None # eg. {"order_number": 12345, "target_qty": 44}

    # final instruction to RPA tool
    rpa_payload: dict[str, Any] | None = None # eg. {"order_number": 12345, "target_qty": 44, "pick_qty_from_location": "WH7",}


@dataclass
class PollResult:
    handled_anything: bool
    handover_data: ActiveJob | None = None



# ============================================================
# EXAMPLE BACKENDS
# ============================================================

# for fetching emails 
class ExampleMailBackend:
    ''' this backend simulates mailbox processing using folders and .eml files. '''
    ''' rewire to eg. Outlook '''


    def __init__(self, log_system, job_source_type) -> None:
        self.log_system = log_system
        self.job_source_type = job_source_type # change to folder in e.g. outlook
        self.inbox_dir = Path(self.job_source_type) / "inbox"
        self.processing_dir = Path(self.job_source_type) / "processing"

        self.inbox_dir.mkdir(parents=True, exist_ok=True)
        self.processing_dir.mkdir(parents=True, exist_ok=True)


    def fetch_from_inbox(self, max_items=None) -> list[str]:
        paths_raw = sorted(self.inbox_dir.glob("*.eml"))

        if max_items is not None:
            paths_raw = paths_raw[:max_items]

        paths = [str(x) for x in paths_raw] #convert Path-type to str

        #self.log_system(f"fetched {paths}")

        return paths
    

    def parse_mail_file(self, processing_path) -> JobCandidate:
        with open(processing_path, "rb") as f:
            msg = BytesParser(policy=policy.default).parse(f)

        from_name, from_address = parseaddr(msg.get("From", ""))
        subject = msg.get("Subject", "").strip()

        del from_name # not used

        #message_id = msg.get("Message-ID", "").strip()
        # not needed. source_ref is sufficient (in this example: Path.   In outlook: Outlook EntryID / Graph ID)

        #raw_headers = {k: str(v) for k, v in msg.items()}   
        # not needed (but good for troubleshooting all metadata) 

        if msg.is_multipart():
            body_parts = []
            for part in msg.walk():
                if part.get_content_type() == "text/plain" and not part.get_filename():
                    try:
                        body_parts.append(part.get_content())
                    except Exception:
                        pass
            body = "\n".join(body_parts).strip()
        else:
            try:
                body = msg.get_content().strip()
            except Exception:
                body = ""
        

        #stub
        attachments = {
            "attachments": [
                {
                    "filename": "orders.xlsx",
                    "path": "/some/path/orders.xlsx",
                }
            ]
        }
     
        return JobCandidate(
            source_ref=processing_path,
            sender_email=from_address.strip().lower(),
            subject=subject,
            body=body,
            job_source_type=self.job_source_type,
            source_data=attachments,
            )



    def claim_to_processing(self, mail: JobCandidate) -> JobCandidate:

        example_path = Path(mail.source_ref) # example backend use Path

        target_path = self.processing_dir / example_path.name #.name gives only the filenamne
        shutil.move(str(example_path), str(target_path))
        
        self.log_system(f"moved {example_path} to {target_path}")
        mail.source_ref = str(target_path)

        return mail
        

    def reply_and_delete(self, candidate: JobCandidate, extra_subject: str, extra_body: str, job_id: int | None = None) -> None:
        self.send_reply(candidate, extra_subject, extra_body, job_id)
        self.delete_from_processing(candidate, job_id)



    def send_reply(self, candidate: JobCandidate, extra_subject: str, extra_body: str, job_id: int | None = None) -> None:
        # DEV STUB

        reply_to = candidate.sender_email
        subject = f"{extra_subject} re: {candidate.subject}"
        body = f"{extra_body} \n\n - - - - - \n{candidate.body}"

        reply_message = f"reply stub to={reply_to}. subject={subject}'. body={body}"
        self.log_system(reply_message[:200], job_id)
        
        print(f"\n*** email reply stub ************\nto={reply_to} \nsubject={subject} \nbody='{body} \n********************************\n")


    def delete_from_processing(self, candidate: JobCandidate, job_id: int | None = None) -> None:

        self.log_system(f"removing: {candidate.source_ref}", job_id)
        os.remove(candidate.source_ref)

    def move_back_to_inbox(self, candidate: JobCandidate) -> JobCandidate:
        ''' to simplify for end-user, return unhandeled emails from shared to origin location'''
        # stub

        # flag/change subject to "FAIL/" ang ignore these in is_shared_inbox_email_in_scope()
        candidate.subject =f"FAIL/ {candidate.subject}" # stub, rename also real email

        example_path = Path(candidate.source_ref)

        target_path = self.inbox_dir / example_path.name #.name only the filenamne
        shutil.move(str(example_path), str(target_path))
        
        self.log_system(f"moved {example_path} back to {target_path}")
        candidate.source_ref = str(target_path)

        return candidate
    
# for query (rewire to eg. real ERP).
class ExampleErpBackend:

    def select_all_from_erp(self, path="Example_ERP_table.xlsx") -> list[dict]:
        # do a well targeted 'query' 

        self.ensure_example_erp_exists(path)

        wb = load_workbook(path)
        ws = wb.active

        assert ws is not None #to satisfy pylance

        all_rows=[]

        for row in ws.iter_rows(min_row=2):  # skip header
            
            source_ref = row[0].value
            order_qty = row[1].value
            material_available = row[2].value

            if order_qty != material_available:

                all_rows.append({
                        "source_ref": source_ref,
                        "order_qty": order_qty,
                        "material_available": material_available,
                    })
                
        wb.close()
        return all_rows
    
    
    def parse_row(self, row) -> JobCandidate:
              
        source_ref = row.get("source_ref")
        order_qty = row.get("order_qty")
        material_available = row.get("material_available")


        try: order_qty = int(order_qty)
        except Exception: raise ValueError(f"invalid order_qty: {order_qty}")
        try: material_available = int(material_available)
        except Exception: raise ValueError(f"invalid material_available: {material_available}")


        source_data ={
            "order_qty": order_qty,
            "material_available": material_available,
        }

        return JobCandidate(
            source_ref=str(source_ref),
            job_source_type="erp_query",
            source_data=source_data
        )

    
    def ensure_example_erp_exists(self, path="Example_ERP_table.xlsx") -> None:
        ''' a table in ERP '''
        if os.path.exists(path):
            return

        wb = Workbook()
        ws = wb.active
        assert ws is not None #to satisfy pylance

        # headers
        ws["A1"] = "source_ref"
        ws["B1"] = "order_qty"
        ws["C1"] = "material_available"

        wb.save(path)
        wb.close()

    
    def get_order_qty(self, source_ref, path="Example_ERP_table.xlsx") -> int | None:
        self.ensure_example_erp_exists(path)

        wb = load_workbook(path)
        ws = wb.active
        assert ws is not None #to satisfy pylance

        for row in ws.iter_rows(min_row=2):
            cell_source_ref = row[0].value

            if str(cell_source_ref) == str(source_ref):
                value = row[1].value  # order_qty    #stype: ignore

                if isinstance(value, int):
                    wb.close()
                    return int(value)
                
                else: 
                    raise ValueError(f"order_qty: {value} is not INT")
        
        wb.close()
        return None  # not found


# ============================================================
# JOB FLOWS
# ============================================================

#for email pipeline
class MailFlow:
    def __init__(self, log_system, log_ui, friends_repo, is_within_operating_hours, network_service, job_handlers, pre_handover_executor) -> None:
        self.log_system = log_system
        self.log_ui = log_ui
        self.friends_repo = friends_repo
        self.is_within_operating_hours = is_within_operating_hours
        self.network_service = network_service
        self.job_handlers = job_handlers
        self.pre_handover_executor = pre_handover_executor
        self.mail_backend_shared = ExampleMailBackend(self.log_system, "shared_inbox",  )
        self.mail_backend_personal = ExampleMailBackend(self.log_system, "personal_inbox",  )

    def poll_once(self) -> PollResult:
        ''' a candidate is an email from personal inbox OR an 'in scope'-email from shared inbox '''

        candidate = self.claim_next_mail_candidate() #claimed and parsed from all mail-sources
        if not candidate:
            return PollResult(handled_anything=False, handover_data=None)
        

        elif candidate.job_source_type == "personal_inbox":
            if self.friends_repo.reload_if_modified():
                self.log_system("friends.xlsx reloaded")

            self.log_ui(f"email from {candidate.sender_email}", blank_line_before=True)
            decision = self.decide_personal_inbox_email(candidate)


        elif candidate.job_source_type == "shared_inbox":
            decision = self.decide_shared_inbox_email(candidate)

        
        else:
            raise RuntimeError(f"unknown source type for candidate: {candidate.job_source_type}")
            
        
        mail_backend = self.get_mail_backend_for_candidate(candidate)
        handover_data = self.pre_handover_executor.execute_decision(candidate, decision, mail_backend)
        return PollResult(handled_anything=True, handover_data=handover_data)


    def claim_next_mail_candidate(self) -> JobCandidate | None:

        # --- personal inbox (parse, always claim) ---
        paths = self.mail_backend_personal.fetch_from_inbox(max_items=1)
       
        for path in paths:
            mail = self.mail_backend_personal.parse_mail_file(path)
            del path
            
            mail = self.mail_backend_personal.claim_to_processing(mail)
            self.log_system(f"{mail.job_source_type} produced mail {mail.source_ref}")
            return mail

        
        # --- shared inbox (parse, maybe claim) ---
        paths = self.mail_backend_shared.fetch_from_inbox()
        
        for path in paths:
            mail = self.mail_backend_shared.parse_mail_file(path)
            del path

            if not self.is_shared_inbox_email_in_scope(mail):
                continue
            
            mail = self.mail_backend_shared.claim_to_processing(mail)
            self.log_system(f"{mail.job_source_type} produced mail {mail.source_ref}")

            return mail


        return None


    def is_shared_inbox_email_in_scope(self, mail: JobCandidate) -> bool:
        #STUB.
        self.log_system(f"checking sender: {mail.sender_email} subject: {mail.subject}")

        # skip emails moved back by move_back_to_inbox()
        if str(mail.subject).upper().startswith("FAIL/"):
            return False
        
        # placeholder for in scope check, eg. invoice from CompanyX that starts with '20....' 
        
        return True

    
    def classify_personal_inbox_email(self, mail: JobCandidate) -> JobType | None:

        subject = str(mail.subject).strip().lower()

        if "ping" in subject.lower():
            return "ping"
        
        elif "job1" in subject.lower():
            return"job1"
        
        elif "job2" in subject.lower():
            return "job2"

        return None




    def classify_shared_inbox_email(self):
        #stub
        pass
 

    def decide_personal_inbox_email(self, mail: JobCandidate) -> JobDecision:
        job_type = None

        try:
            if not self.friends_repo.is_allowed_sender(mail.sender_email):
                return JobDecision(
                    action="DELETE_ONLY",
                    ui_log_message="--> rejected (not in friends.xlsx)",
                    system_log_message="--> rejected (not in friends.xlsx)"
                    
                )

            if not self.is_within_operating_hours():
                return JobDecision(
                    action="REPLY_AND_DELETE",
                    job_status="REJECTED",
                    error_code="OUTSIDE_WORKING_HOURS",
                    error_message="Email received outside working hours 05-23.",
                    ui_log_message="--> rejected (outside working hours)",
                    system_log_message="--> rejected (outside working hours)",
                )

            job_type = self.classify_personal_inbox_email(mail)

            if job_type == None:
                return JobDecision(
                    action="REPLY_AND_DELETE",
                    job_type=None,
                    job_status="REJECTED",
                    error_code="UNKNOWN_JOB",
                    error_message="Could not identify a job type.",
                    ui_log_message="--> rejected (unable to identify job type)",
                    system_log_message="--> rejected (unable to identify job type)",
                )
            

            if not self.friends_repo.has_job_access(mail.sender_email, job_type):
                return JobDecision(
                    action="REPLY_AND_DELETE",
                    job_type=job_type,
                    job_status="REJECTED",
                    error_code="NO_ACCESS",
                    error_message=f"No access to {job_type}. Check with administrator for access.",
                    ui_log_message=f"--> rejected (no access to {job_type})",
                    system_log_message=f"--> rejected (no access to {job_type})",
                )



            if not self.network_service.has_network_access():
                return JobDecision(
                    action="REPLY_AND_DELETE",
                    job_type=job_type,
                    job_status="REJECTED",
                    error_code="NO_NETWORK",
                    error_message="No network connection. Your email was removed.",
                    ui_log_message="--> rejected (no network connection)",
                    system_log_message="--> rejected (no network connection)",
                )

            handler = self.job_handlers.get(job_type)
            if handler is None:
                return JobDecision(
                    action="CRASH",
                    job_type=job_type,
                    error_message=f"No handler registered for job_type={job_type}",
                )

            ok, payload_or_error = handler.precheck_and_build_executiondata(mail)
            if not ok:
                error = str(payload_or_error)
                return JobDecision(
                    action="REPLY_AND_DELETE",
                    job_type=job_type,
                    job_status="REJECTED",
                    error_code="INVALID_INPUT",
                    error_message=error,
                    ui_log_message=f"--> rejected (invalid input for {job_type})",
                    system_log_message=f"--> rejected (invalid input for {job_type})",
                )
            
            
            rpa_payload = payload_or_error

            return JobDecision(
                action="QUEUE_RPA_JOB",
                job_type=job_type,
                job_status="QUEUED",
                system_log_message=f"accepted ({job_type})",
                send_lifesign_notice=True,
                start_recording=True,
                rpa_payload=rpa_payload
            )

        except Exception as err:
            return JobDecision(
                action="CRASH",
                job_type=None,
                error_message=str(err),
            )
    
 
    def decide_shared_inbox_email(self, mail: JobCandidate) -> JobDecision:
        #stub
        return JobDecision(
                    action="MOVE_BACK_TO_INBOX",
                    system_log_message=f"No logic yet, move back this email to inbox from proccessing folder: {mail.sender_email}" #only in DEV
                )



    def get_mail_backend_for_candidate(self, mail: JobCandidate) -> ExampleMailBackend:
        if mail.job_source_type == "personal_inbox":
            return self.mail_backend_personal
        if mail.job_source_type == "shared_inbox":
            return self.mail_backend_shared
        raise ValueError(f"unknown job_source_type={mail.job_source_type}")

# for scheduledjobs pipeline
class ScheduledFlow:
    ''' scheduled jobs pipeline '''
    def __init__(self, log_system, log_ui, audit_repo, job_handlers, in_dev_mode, pre_handover_executor) -> None:
        self.in_dev_mode = in_dev_mode
        self.log_system = log_system
        self.log_ui = log_ui
        self.audit_repo = audit_repo
        self.job_handlers = job_handlers
        self.pre_handover_executor = pre_handover_executor
        self.erp_backend = ExampleErpBackend()

        self.poll_interval = 1 if self.in_dev_mode else 600   # 600 = 10 min
        self.next_check_time = 0


    def poll_once(self) -> PollResult:
        #candidate can be a row from a query 

        now = time.time()

        if now > self.next_check_time:
            self.next_check_time = now + self.poll_interval

            candidate = self.fetch_next_scheduled_candidate()
            if not candidate:
                return PollResult(handled_anything=False, handover_data=None)


            self.log_ui(f"scheduled job detected: {candidate.source_ref}", blank_line_before=True)
            decision = self.decide_candidate(candidate)

            
            handover_data = self.pre_handover_executor.execute_decision(candidate, decision)
            return PollResult(handled_anything=True, handover_data=handover_data)
        
        return PollResult(handled_anything=False, handover_data=None)


    def fetch_next_scheduled_candidate(self) -> JobCandidate | None:

        # job 3
        all_selected_rows_query3 = self.erp_backend.select_all_from_erp()
        
        if not all_selected_rows_query3:
            return None
    
        for row_candidate_raw in all_selected_rows_query3:
            row_candidate = self.erp_backend.parse_row(row_candidate_raw)

            # avoid bad loops by not working the same row twice a day
            if self.audit_repo.has_been_processed_today(row_candidate.source_ref):
                continue

            row_candidate.job_source_type="erp_query"
            self.log_system(f"{row_candidate.job_source_type} produced source_ref {row_candidate.source_ref}")
            return row_candidate
        
        # job 4
        # stub
        
        return None


    def decide_candidate(self, candidate_row: JobCandidate) -> JobDecision:
        self.log_system("running")

        job_type = None

        try:

                        #placeholder evaluation logic
                        # eg. below:


            job_type = self.classify_candidate(candidate_row)
            handler = self.job_handlers.get(job_type)

            if handler is None:
                return JobDecision(
                    action="CRASH",
                    job_type=job_type,
                    error_message=f"No handler found for job_type={job_type}",
                )
            
            # do the precheck from "job specifics"-section
            ok, payload_or_error = handler.precheck_and_build_executiondata(candidate_row)

            if not ok:
                error = str(payload_or_error)
                return JobDecision(
                    action="SKIP",  
                    job_type=job_type,
                    job_status="REJECTED",
                    error_code="INVALID_INPUT",
                    error_message=error,
                    ui_log_message=f"--> rejected (invalid input for {job_type})",
                )

            rpa_payload = payload_or_error
            
            
            return JobDecision(
                action="QUEUE_RPA_JOB",
                job_type=job_type,
                job_status="QUEUED",
                system_log_message=f"accepted ({job_type})",
                start_recording=True,
                rpa_payload=rpa_payload,
                )

        except Exception as err:
            return JobDecision(
                action="CRASH",
                job_type=job_type,
                error_message=str(err),
            )
    

    def classify_candidate(self, candidate: JobCandidate) -> JobType | None:
        #stub
        del candidate
        self.log_system("STUB. running: 'job3'")
        return "job3"

# for decision-making on the found job
class PreHandoverExecutor:
    def __init__(self, log_system, log_ui, update_ui_status, ui_dot_tk_set_show_recording_overlay, generate_job_id, recording_service, audit_repo, in_dev_mode) -> None:
        self.in_dev_mode = in_dev_mode
        self.log_system = log_system
        self.log_ui = log_ui
        self.recording_service = recording_service
        self.generate_job_id = generate_job_id
        self.audit_repo = audit_repo
        self.update_ui_status = update_ui_status
        self.ui_dot_tk_set_show_recording_overlay = ui_dot_tk_set_show_recording_overlay
    
    def validate_decision_itself(self, decision: JobDecision) -> None:

        if not isinstance(decision, JobDecision):
            raise ValueError("decision must be JobDecision")

        # --- validate action ---
        if decision.action not in get_args(JobAction):
            raise ValueError(f"invalid action: {decision.action}")

        # --- validate job_type ---
        if decision.job_type is not None and decision.job_type not in get_args(JobType):
            raise ValueError(f"invalid job_type: {decision.job_type}")

        # --- validate job_status ---
        if decision.job_status is not None and decision.job_status not in get_args(JobStatus):
            raise ValueError(f"invalid job_status: {decision.job_status}")

        # ============================================================
        # ACTION-SPECIFIC RULES
        # ============================================================

        action = decision.action

        # ------------------------------------------------------------
        # DELETE_ONLY
        # ------------------------------------------------------------
        if action == "DELETE_ONLY":
            if decision.job_status is not None:
                raise ValueError("DELETE_ONLY must not have job_status")
            if decision.rpa_payload is not None:
                raise ValueError("DELETE_ONLY must not have rpa_payload")

        # ------------------------------------------------------------
        # REPLY_AND_DELETE
        # ------------------------------------------------------------
        elif action == "REPLY_AND_DELETE":
            if decision.job_status != "REJECTED":
                raise ValueError("REPLY_AND_DELETE requires job_status REJECTED")

            if decision.error_message is None:
                raise ValueError("REPLY_AND_DELETE requires error_message (reply text)")

            if decision.rpa_payload is not None:
                raise ValueError("REPLY_AND_DELETE must not have rpa_payload")

        # ------------------------------------------------------------
        # QUEUE_RPA_JOB
        # ------------------------------------------------------------
        elif action == "QUEUE_RPA_JOB":
            if decision.job_type is None:
                raise ValueError("QUEUE_RPA_JOB requires job_type")

            if decision.job_status != "QUEUED":
                raise ValueError("QUEUE_RPA_JOB requires job_status='QUEUED'")

            if decision.rpa_payload is None:
                raise ValueError("QUEUE_RPA_JOB requires rpa_payload")

            if not isinstance(decision.rpa_payload, dict):
                raise ValueError("rpa_payload must be dict")

        # ------------------------------------------------------------
        # SKIP (scheduled flow reject)
        # ------------------------------------------------------------
        elif action == "SKIP":
            if decision.job_status != "REJECTED":
                raise ValueError("SKIP requires job_status='REJECTED'")

        # ------------------------------------------------------------
        # MOVE_BACK_TO_INBOX
        # ------------------------------------------------------------
        elif action == "MOVE_BACK_TO_INBOX":
            if decision.job_status is not None:
                raise ValueError("MOVE_BACK_TO_INBOX should not set job_status")

        # ------------------------------------------------------------
        # CRASH
        # ------------------------------------------------------------
        elif action == "CRASH":
            if decision.error_message is None:
                raise ValueError("CRASH requires error_message")

        else:
            # should never happen due to earlier validation
            raise ValueError(f"Unhandled action: {action}")

        # ============================================================
        # GENERIC SANITY CHECKS
        # ============================================================

        # UI / system messages should be strings if present
        if decision.ui_log_message is not None and not isinstance(decision.ui_log_message, str):
            raise ValueError("ui_log_message must be str")

        if decision.system_log_message is not None and not isinstance(decision.system_log_message, str):
            raise ValueError("system_log_message must be str")

        # flags
        if not isinstance(decision.send_lifesign_notice, bool):
            raise ValueError("send_lifesign_notice must be bool")

        if not isinstance(decision.start_recording, bool):
            raise ValueError("start_recording must be bool")
        

    def validate_candidate_decision_combination(self, candidate: JobCandidate, decision: JobDecision) -> None:

        if not isinstance(candidate, JobCandidate):
            raise ValueError("candidate must be JobCandidate")

        if candidate.job_source_type not in get_args(JobSourceType):
            raise ValueError(f"invalid candidate.job_source_type: {candidate.job_source_type}")

        is_mail = candidate.job_source_type in ("personal_inbox", "shared_inbox")
        is_scheduled = candidate.job_source_type == "erp_query"

        # ------------------------------------------------------------
        # source-type-specific candidate sanity
        # ------------------------------------------------------------
        if is_mail:
            if candidate.sender_email is None:
                raise ValueError(f"{candidate.job_source_type} candidate requires sender_email")
            if candidate.subject is None:
                raise ValueError(f"{candidate.job_source_type} candidate requires subject")
            if candidate.body is None:
                raise ValueError(f"{candidate.job_source_type} candidate requires body")

        elif is_scheduled:
            if candidate.source_data is None:
                raise ValueError("erp_query candidate requires source_data")

        else:
            raise ValueError(f"unknown candidate.job_source_type: {candidate.job_source_type}")

        # ------------------------------------------------------------
        # action must match candidate source type
        # ------------------------------------------------------------
        if decision.action in ("DELETE_ONLY", "REPLY_AND_DELETE", "MOVE_BACK_TO_INBOX"):
            if not is_mail:
                raise ValueError(
                    f"action {decision.action} is only valid for mail candidates, "
                    f"not {candidate.job_source_type}"
                )

        if decision.action == "SKIP":
            if not is_scheduled:
                raise ValueError(
                    f"action SKIP is only valid for scheduled candidates, "
                    f"not {candidate.job_source_type}"
                )

        # ------------------------------------------------------------
        # optional policy checks
        # ------------------------------------------------------------
        if decision.send_lifesign_notice and not is_mail:
            raise ValueError("send_lifesign_notice is only valid for mail candidates")

        # MOVE_BACK_TO_INBOX is for shared-only emails
        if decision.action == "MOVE_BACK_TO_INBOX":
            if candidate.job_source_type != "shared_inbox":
                raise ValueError("MOVE_BACK_TO_INBOX is only valid for shared_inbox")

        # DELETE_ONLY is for personal-only emails.
        if decision.action == "DELETE_ONLY":
            if candidate.job_source_type != "personal_inbox":
                raise ValueError("DELETE_ONLY is only valid for personal_inbox")


        # REPLY_AND_DELETE should only be used for personal inbox only 
        if decision.action == "REPLY_AND_DELETE" and candidate.job_source_type != "personal_inbox":
            raise ValueError("REPLY_AND_DELETE is only valid for personal_inbox")

        # send_lifesign_notice only makes sense together with QUEUE_RPA_JOB
        if decision.send_lifesign_notice and decision.action != "QUEUE_RPA_JOB":
            raise ValueError("send_lifesign_notice requires action='QUEUE_RPA_JOB'")

        # Optional stricter policy:
        # start_recording only makes sense for queued jobs
        if decision.start_recording and decision.action != "QUEUE_RPA_JOB":
            raise ValueError("start_recording requires action='QUEUE_RPA_JOB'")
        

    # standard workflow is that executor delegates work to RPA tool
    def execute_decision(self, candidate: JobCandidate, decision: JobDecision, mail_backend: ExampleMailBackend | None=None) -> ActiveJob | None:
        
        self.validate_decision_itself(decision)
        self.validate_candidate_decision_combination(candidate, decision)

        is_mail = candidate.job_source_type in ("personal_inbox", "shared_inbox")
        is_scheduled = candidate.job_source_type == "erp_query"

        if decision.ui_log_message:
            self.log_ui(decision.ui_log_message)

        if decision.system_log_message:
            self.log_system(decision.system_log_message)

        # ------------------------------------------------------------
        # mail actions
        # ------------------------------------------------------------

        if is_mail:
            if mail_backend is None:
                raise ValueError("mail_backend required for mail actions")

            if decision.action == "MOVE_BACK_TO_INBOX": # eg. error with emails in scope from shared inbox
                mail_backend.move_back_to_inbox(candidate)
                return None

            if decision.action == "DELETE_ONLY":
                mail_backend.delete_from_processing(candidate) 
                return None
            
            #"PONG (robot online)."
            if decision.action == "REPLY_AND_DELETE":
                job_id = self.generate_job_id()
                now = datetime.datetime.now()

                assert decision.error_message is not None #to satisfy pylance
                
                mail_backend.reply_and_delete(
                    candidate,
                    extra_subject="FAIL",
                    extra_body=decision.error_message,
                    job_id=job_id,
                    )

                self.audit_repo.insert_job(
                    job_id=job_id,
                    email_address=candidate.sender_email,
                    email_subject=candidate.subject,
                    job_type=decision.job_type,
                    job_start_date=now.strftime("%Y-%m-%d"),
                    job_start_time=now.strftime("%H:%M:%S"),
                    job_finish_time=now.strftime("%H:%M:%S"),
                    job_status=decision.job_status,
                    job_source_type=candidate.job_source_type,
                    final_reply_sent=True,
                    error_code=decision.error_code,
                    error_message=decision.error_message,
                )

                return None

        # ------------------------------------------------------------
        # scheduled-only actions
        # ------------------------------------------------------------
        if is_scheduled:
            if decision.action == "SKIP":
                job_id = self.generate_job_id()
                now = datetime.datetime.now()

                self.audit_repo.insert_job(
                    job_id=job_id,
                    source_ref=candidate.source_ref,
                    job_type=decision.job_type,
                    job_start_date=now.strftime("%Y-%m-%d"),
                    job_start_time=now.strftime("%H:%M:%S"),
                    job_finish_time=now.strftime("%H:%M:%S"),
                    job_status=decision.job_status,
                    job_source_type=candidate.job_source_type,
                    error_code=decision.error_code,
                    error_message=decision.error_message,
                )
                return None
            
        # ------------------------------------------------------------
        # queue RPA job
        # ------------------------------------------------------------
        if decision.action == "QUEUE_RPA_JOB":
            job_id = self.generate_job_id()
            now = datetime.datetime.now()
            
            self.update_ui_status(forced_status="working")

            if is_mail:
                assert mail_backend is not None # to satisfy pylance
                
                if decision.send_lifesign_notice and not self.audit_repo.has_sender_job_today(candidate.sender_email, job_id):
                    mail_backend.send_reply(
                        candidate=candidate,
                        extra_subject="ONLINE",
                        extra_body = (">HELLO HUMAN\n\n"
                        "This is an automated system reply.\n\n"
                        "It appears to be your first request today, so this reply confirms that the robot is online.\n"
                        "Your job has been received and is now processing.\n"
                        "You will receive another message when the job is completed."),
                        job_id=job_id,
                    )

            self.audit_repo.insert_job(
                job_id=job_id,
                source_ref=candidate.source_ref,
                email_address=candidate.sender_email if is_mail else None,
                email_subject=candidate.subject if is_mail else None,
                job_type=decision.job_type,
                job_start_date=now.strftime("%Y-%m-%d"),
                job_start_time=now.strftime("%H:%M:%S"),
                job_status="QUEUED",
                job_source_type=candidate.job_source_type,
            )

            if decision.start_recording:
                if not self.in_dev_mode:
                    self.recording_service.start(job_id)
                self.ui_dot_tk_set_show_recording_overlay()
            
            return ActiveJob(
                ipc_state="job_queued",
                job_id=job_id,
                job_type=decision.job_type,
                job_source_type=candidate.job_source_type,
                source_ref=candidate.source_ref,
                sender_email=candidate.sender_email,
                subject=candidate.subject,
                body=candidate.body,
                source_data=candidate.source_data,
                rpa_payload=decision.rpa_payload,
                )

        # ------------------------------------------------------------
        # crash
        # ------------------------------------------------------------
        if decision.action == "CRASH":
            job_id = self.generate_job_id()
            now = datetime.datetime.now()
            final_reply_flag = False

            if candidate.job_source_type == "personal_inbox":
                try:                    
                    if mail_backend is None: raise ValueError("mail_backend is None")
                    mail_backend.reply_and_delete(
                        candidate,
                        extra_subject="FAIL",
                        extra_body="System crash, the robot is out-of-service and your email was deleted.",
                        job_id=job_id,
                    )
                    final_reply_flag = True
                except Exception:
                    try: self.log_system("WARN: unable to notify user of crash", job_id)
                    except Exception: print(f"WARN: unable to notify user of crash, jobid {job_id}")

            try:
                self.audit_repo.insert_job(
                    job_id=job_id,
                    source_ref=candidate.source_ref,
                    email_address=candidate.sender_email if is_mail else None,
                    email_subject=candidate.subject if is_mail else None,
                    job_type=decision.job_type,
                    job_start_date=now.strftime("%Y-%m-%d"),
                    job_start_time=now.strftime("%H:%M:%S"),
                    job_finish_time=now.strftime("%H:%M:%S"),
                    job_status="FAIL",
                    job_source_type=candidate.job_source_type,
                    final_reply_sent = final_reply_flag,
                    error_code="SYSTEM_CRASH",
                    error_message=decision.error_message,
                )
            except Exception:
                try: self.log_system("WARN: unable to update db", job_id)
                except Exception: print(f"WARN: unable to update db, jobid {job_id}")
                

            self.log_ui("--> rejected (system crash)")
            raise RuntimeError(f"job_id{job_id} crashed: {decision.error_message}") # policy safe-stop for all unexpected
            

        raise ValueError(f"decision.action={decision.action} is not valid for specified candidate type") # policy safe-stop for all unexpected

# for closing the job
class PostHandoverFinalizer:
    ''' the verification step, if any, is always a cold start '''
    def __init__(self, log_system, log_ui, audit_repo, job_handlers, recording_service, ui_dot_tk_set_hide_recording_overlay, refresh_jobs_done_today_display, in_dev_mode) -> None:
        self.in_dev_mode = in_dev_mode

        self.log_system = log_system
        self.log_ui = log_ui
        self.audit_repo = audit_repo
        self.job_handlers = job_handlers
        self.recording_service = recording_service
        self.ui_dot_tk_set_hide_recording_overlay = ui_dot_tk_set_hide_recording_overlay
        self.refresh_jobs_done_today_display = refresh_jobs_done_today_display
        self.mail_backend_personal = ExampleMailBackend(self.log_system, "personal_inbox",  )
        self.mail_backend_shared = ExampleMailBackend(self.log_system, "shared_inbox",  )


    def poll_once(self, handover_data) -> None:
        # rebuild active_job objekt after cold start
        active_job = self.rebuild_active_job(handover_data)

        #get id and type
        job_id = active_job.job_id
        job_type = active_job.job_type

        # note in audit that the job is 'taken back' from RPA
        self.log_system(f"fetched: {handover_data}", job_id)
        self.audit_repo.update_job(
            job_id=job_id,
            job_status="VERIFYING"
            )

        # use job-specific verfification 
        handler = self.job_handlers.get(job_type)
        if handler is None:
            ok_or_error = f"No handler for job_type={job_type}"

        else:
            try:
                ok_or_error = handler.verify_result(active_job)
            except Exception as err:
                ok_or_error = f"verification crash: {err}"

        ok_or_error = self.finalize_job_result(ok_or_error, active_job)

        return ok_or_error

    

    def finalize_job_result(self, ok_or_error, active_job: ActiveJob):

        job_status: JobStatus

        if ok_or_error == "ok":
            job_status = "DONE"
            error_code = None
            error_message = None
        else:
            job_status = "FAIL"
            error_message = ok_or_error
            error_code="VERIFICATION_FAIL"

        
        job_id = active_job.job_id
        job_type = active_job.job_type

        self.recording_service.stop(job_id) 
        self.ui_dot_tk_set_hide_recording_overlay() # the " *RECORDING "-box

        if not self.in_dev_mode: self.recording_service.upload_recording(job_id=job_id)

 
        # for source specific (email or erp)
        final_reply_sent = self.handle_source_completion(active_job, job_status, error_message)

        # update ui log
        self.log_ui(f"--> {job_status.lower()} ({job_type})")

        # update audit w/ result (DONE/FAIL)
        self.audit_repo.update_job(
            job_id=job_id, 
            job_status=job_status, 
            error_code=error_code, 
            error_message=error_message, 
            job_finish_time=datetime.datetime.now().strftime("%H:%M:%S"),
            final_reply_sent=final_reply_sent,
            )
        

        self.refresh_jobs_done_today_display()

        # policy to safe-stop if verification failed
        if ok_or_error != "ok":
            raise RuntimeError(f"job_id {job_id} crashed, verification failed: {ok_or_error}") 


    def create_reply_body(self, active_job: ActiveJob, job_status: str, error_message: str | None, ) -> str:
        job_id = active_job.job_id

        recording_path = self.get_recording_path_for_user(active_job)

        recording_text = ""
        if recording_path:
            recording_text = (f"A screen recording is available:\n {recording_path} \n If the link is not clickable, copy the path and open it from File Explorer."
            )

        if job_status == "DONE":
            return (
                f"Your request has been completed successfully.\n\n"
                f"Job ID: {job_id}\n\n"
                f"{recording_text}\n\n"
                f"This email can be deleted."
            )

        if job_status == "FAIL":
            reason = self.make_user_safe_error_text(error_message)

            return (
                f"Your request could not be completed.\n\n"
                f"Job ID: {job_id}\n"
                f"Reason: {reason}\n\n"
                f"{recording_text}\n\n"
                f"To avoid further problems, the robot will now go out-of-service.\n"
                f"No action is required from your side.\n\n"
                f"This email can be deleted."
            )

        raise ValueError(f"unsupported job_status: {job_status}")


    def make_user_safe_error_text(self, error_message: str | None) -> str:
        if not error_message:
            return "The job failed during processing."

        text = str(error_message).strip()
        lowered = text.lower()

        if "timeout" in lowered:
            return "The automation took too long and was stopped."

        if "verification" in lowered:
            return "The result could not be verified after execution."

        if "network" in lowered:
            return "A network-related problem occurred during processing."

        if "missing rpa_payload" in lowered:
            return "The job could not be prepared correctly for execution."

        if "crash" in lowered:
            return "The automation encountered a system error."

        return text[:300]

        

        
    def handle_source_completion(self, active_job: ActiveJob, job_status: str, error_message: str | None) -> bool:
        # for erp
        if active_job.job_source_type == "erp_query":
            return False

        # #for email
        if active_job.source_ref is None:
            raise ValueError("missing source_ref")
        if active_job.source_data is None:
            raise ValueError("missing source_data")

        if active_job.job_source_type in ("personal_inbox", "shared_inbox"):
            candidate = JobCandidate(
                source_ref=active_job.source_ref,
                job_source_type=active_job.job_source_type,
                source_data=active_job.source_data,
                sender_email=active_job.sender_email,
                subject=active_job.subject,
                body=active_job.body,
                )


        if active_job.job_source_type == "personal_inbox":
            reply_body = self.create_reply_body(active_job, job_status, error_message)

            self.mail_backend_personal.reply_and_delete(
                candidate=candidate,
                job_id=active_job.job_id,
                extra_subject=job_status,
                extra_body=reply_body,
            )
            return True

        # delete shared mail (or move to archive?)
        if active_job.job_source_type == "shared_inbox":
            self.mail_backend_shared.delete_from_processing(candidate, job_id=active_job.job_id)
            return False

        raise ValueError(f"unknown job_source_type={active_job.job_source_type}")
        

    def rebuild_active_job(self, handover_data: dict) -> ActiveJob:
        if not isinstance(handover_data, dict):
            raise ValueError("handover_data must be dict")

        ipc_state = handover_data.get("ipc_state")
        if ipc_state not in get_args(IpcState):
            raise ValueError(f"invalid ipc_state: {ipc_state}")
        
        assert ipc_state is not None # to satisfy pylance

        job_type = handover_data.get("job_type")
        if job_type is not None and job_type not in get_args(JobType):
            raise ValueError(f"invalid job_type: {job_type}")

        job_source_type = handover_data.get("job_source_type")
        if job_source_type is not None and job_source_type not in get_args(JobSourceType):
            raise ValueError(f"invalid job_source_type: {job_source_type}")

        job_id = handover_data.get("job_id")
        if job_id is not None:
            try:
                job_id = int(job_id)
            except Exception:
                raise ValueError(f"job_id not int-like: {job_id}")

        source_data = handover_data.get("source_data") or {}
        rpa_payload = handover_data.get("rpa_payload") or {}

        return ActiveJob(
            ipc_state=ipc_state,
            source_ref=handover_data.get("source_ref"),
            job_type=job_type,
            job_source_type=job_source_type,
            job_id=job_id,
            sender_email=handover_data.get("sender_email"),
            subject=handover_data.get("subject"),
            body=handover_data.get("body"),
            source_data=source_data,
            rpa_payload=rpa_payload,
        )


    def get_mail_backend_for_candidate(self, candidate: JobCandidate) -> ExampleMailBackend:
        if candidate.job_source_type == "personal_inbox":
            return self.mail_backend_personal
        if candidate.job_source_type == "shared_inbox":
            return self.mail_backend_shared
        raise ValueError(f"unknown email job_source_type={candidate.job_source_type}")
        

    def get_recording_path_for_user(self, active_job: ActiveJob) -> str | None:
        job_id = active_job.job_id
        if job_id is None:
            return None

        network_path = Path(RecordingService.RECORDINGS_DESTINATION_FOLDER) / f"{job_id}.mkv" #replce with below if on shared drive
        #network_path = Path(r"\\server\recordings") / f"{job_id}.mkv"

        if network_path.exists():
            return str(network_path)

        return None


# ============================================================
# JOB SPECIFICS
# ============================================================

# for job1 (inbox jobsource)
class ExampleJob1Handler:
    ''' everything for job1 '''
    def __init__(self,log_system) -> None:
        self.log_system = log_system

    # sanity-check (and ERP check) on given data
    def precheck_and_build_executiondata(self, candidate: JobCandidate) -> tuple[bool, dict | str]:
        body = candidate.body
        assert body is not None # to satisfy pylance


        # get important info for job1, eg.:
        order_number_match = re.search(r"order_number:\s*(.+)", body)
        order_number = order_number_match.group(1) if order_number_match else None

        order_qty_match = re.search(r"order_qty:\s*(.+)", body)
        order_qty = order_qty_match.group(1) if order_qty_match else None

        material_available_match = re.search(r"material_available:\s*(.+)", body)
        material_available = material_available_match.group(1) if material_available_match else None

        error = ""
        if order_number is None:
            error += "missing source_ref. "
        if order_qty is None:
            error += "missing order_qty. "
        if material_available is None:
            error += "missing material_available. "

        if error:
            return False, error.strip()

        # and for any attachments, eg:
        attachments = candidate.source_data.get("attachments", [])
        #for attachment in attachments:
        #    print(attachment.get("filename"))

        rpa_payload = {
            "source_ref": order_number,
            "order_qty": order_qty,
            "target_order_qty": material_available,
            "attachments": attachments,
        }

        return True, rpa_payload
    

    def verify_result(self, activejob: ActiveJob):
        return "ok"

# for job2 (inbox jobsource)
class ExampleJob2Handler:
    ''' everything for job2 '''
    def __init__(self,log_system) -> None:
        self.log_system = log_system
   
    def precheck_and_build_executiondata(self, candidate: JobCandidate) -> tuple[bool, dict | str]:
        return False, "stub. no logic for job2."

    def verify_result(self, activejob: ActiveJob):
        return "ok"

# for ping (inbox jobsource)
class ExamplePingJobHandler:
    ''' everything for ping '''
    def __init__(self,log_system) -> None:
        self.log_system = log_system

    def precheck_and_build_executiondata(self, candidate: JobCandidate) -> tuple[bool, dict | str]:
        return True, {}
    
    def verify_result(self, activejob: ActiveJob):
        return "ok"
    
   
# for job3 (ERP jobsource)
class ExampleJob3Handler:
    ''' everything for job3 '''
    def __init__(self, log_system) -> None:
        self.log_system = log_system
        self.erp_backend = ExampleErpBackend()

   
    def precheck_and_build_executiondata(self, candidate: JobCandidate) -> tuple[bool, dict[str, Any] | str]:
        source_ref = candidate.source_ref
        order_qty = candidate.source_data.get("order_qty")
        material_available = candidate.source_data.get("material_available")

        if order_qty == material_available:
            return False, "no mismatch left to fix"

        rpa_payload = {
            "source_ref": str(source_ref),
            "target_order_qty": material_available,
        }

        return True, rpa_payload
    

    def verify_result(self, activejob: ActiveJob) -> str:
        
        # get erp order numer/id
        rpa_payload = activejob.rpa_payload
        if not rpa_payload:
            return "missing rpa_payload"
        
         # get the order number/id and the target qty sent to RPA tool
        source_ref = rpa_payload.get("source_ref")
        target_order_qty = rpa_payload.get("target_order_qty")


        # get the 'real' qty now in erp
        order_qty_erp = self.erp_backend.get_order_qty(source_ref)

        # compare them
        if order_qty_erp != target_order_qty:
            message= f"ERP still shows mismatch after RPA update. Should be: {target_order_qty}, is: {order_qty_erp}"
            self.log_system(message, activejob.job_id)
            return message

        self.log_system(f"OK. Should be: {target_order_qty}, is: {order_qty_erp}", activejob.job_id)
        return "ok"


# ============================================================
# HANDOVER / IPC
# ============================================================

# for file IPC
class HandoverRepository:
    ''' HANDOVER_FILE is the I/O between this script and the RPA tool  '''

    def __init__(self, log_system) -> None:
        self.log_system = log_system
        self.HANDOVER_FILE = "handover.json"


    def read(self) -> dict:
        ''' read HANDOVER_FILE '''
        
        last_err=None

        for attempt in range(7):
            try:
                with open(self.HANDOVER_FILE, "r", encoding="utf-8") as f:
                    handover_data = json.load(f)

                self.validate_handover_data(handover_data)
                return handover_data

            except Exception as err:
                last_err = err
                print(f"WARN: retry {attempt+1}/7 : {err}")
                #time.sleep((attempt+1) ** 2) fail fast in dev
        
        
        raise RuntimeError(f"{self.HANDOVER_FILE} unreadable: {last_err}")
    
      
    def write(self, handover_data: ActiveJob) -> None:
        ''' atomic write of HANDOVER_FILE '''

        handover_data_asdict = asdict(handover_data)
        self.validate_handover_data(handover_data_asdict)

        for attempt in range(7):
            temp_path = None
            try:
                dir_path = os.path.dirname(os.path.abspath(self.HANDOVER_FILE))
                fd, temp_path = tempfile.mkstemp(dir=dir_path, suffix=".tmp")    # create temp file

                #atomic write
                with os.fdopen(fd, "w", encoding="utf-8") as tmp:
                    json.dump(handover_data_asdict, tmp, indent=2) # indent for human eyes
                    tmp.flush()
                    os.fsync(tmp.fileno())

                os.replace(temp_path, self.HANDOVER_FILE) # replace original file
                self.log_system(f"written: {handover_data_asdict}", job_id=handover_data_asdict.get("job_id"))
                return

            except Exception as err:
                last_err = err
                print(f"{attempt+1}st warning from write()")
                self.log_system(f"WARN: {attempt+1}/7 error", job_id=handover_data_asdict.get("job_id"))
                #time.sleep(attempt + 1) # 1 2... 7 sec       #fail fast in dev

            finally: #remove temp-file if writing fails.
                if temp_path and os.path.exists(temp_path):
                    try: os.remove(temp_path)
                    except Exception: pass

        self.log_system(f"CRITICAL: cannot write {self.HANDOVER_FILE} {last_err}", job_id=handover_data_asdict.get("job_id"))
        raise RuntimeError(f"CRITICAL: cannot write {self.HANDOVER_FILE}")
  

    def validate_handover_data(self, handover_data: dict): 
        ''' check some basic combinations '''

        ipc_state = handover_data.get("ipc_state")
        job_id = handover_data.get("job_id")
        job_type = handover_data.get("job_type")
        job_source_type = handover_data.get("job_source_type")
        source_ref = handover_data.get("source_ref") 
        sender_email = handover_data.get("sender_email")
        subject = handover_data.get("subject")
        body = handover_data.get("body")
        source_data = handover_data.get("source_data")
        rpa_payload = handover_data.get("rpa_payload")


        # common fields checks
        if job_id is not None:
            try: job_id = int(job_id)
            except Exception: raise ValueError(f"job_id not INT-like: {job_id}")
        
        if ipc_state == None:
            raise ValueError(f"ipc_state missing")
        
        if ipc_state == "idle" and (job_id or job_type or job_source_type or source_ref or sender_email or subject or body or source_data or rpa_payload):
            raise ValueError(f"state 'idle' should have no more variables: {handover_data}")

        if ipc_state not in get_args(IpcState):
            raise ValueError(f"unknown state: {ipc_state}")
        
        if ipc_state in ("job_queued", "job_running", "job_verifying"):
   
            required_fields = {
            "job_id": job_id,
            "job_type": job_type,
            "source_ref": source_ref,
            "job_source_type": job_source_type,
            "rpa_payload": rpa_payload,
            }

            missing = [k for k, v in required_fields.items() if v is None]  # allowing value "0"
            if missing:
                raise ValueError(f"{ipc_state} has missing fields in {self.HANDOVER_FILE}: {missing}")
            
            if job_type not in get_args(JobType):
                raise ValueError(f"unkown job_type: {job_type} for {ipc_state}")

         
            if job_source_type not in get_args(JobSourceType):
                raise ValueError(f"unknown job_source_type: {job_source_type}")
           
            
            # mail specific checks
            if job_source_type in ("personal_inbox", "shared_inbox"):
                
                required_fields = {
                "source_ref": source_ref,
                "sender_email": sender_email,
                "subject": subject,
                "body": body,
                }
            
            # scheduled specific checks
            elif job_source_type == "erp_query":

                required_fields = {
                "source_ref": source_ref,
                "source_data": source_data,
                }

            missing = [k for k, v in required_fields.items() if v is None] # allowing value "0"
            if missing:
                raise ValueError(f"{job_source_type} has missing fields in {self.HANDOVER_FILE}: {missing}")
            

# ============================================================
# RECORDING / SAFESTOP / INFRASTRUCTURE
# ============================================================
                          
# for screen recording
class RecordingService:
    ''' screen-recording to capture all RPA tool screen-activity '''

    RECORDINGS_IN_PROGRESS_FOLDER = "recordings_in_progress"
    RECORDINGS_DESTINATION_FOLDER = "recordings_destination"

    def __init__(self, log_system,) -> None:
        

        self.log_system = log_system
        self.recording_process = None

    #start the recording
    def start(self, job_id) -> None:

        if platform.system() == "Windows" and not os.path.exists("./ffmpeg.exe"):
            message ="WARN: screen-recording disabled due to missing file ffmpeg.exe, download from eg. https://www.gyan.dev/ffmpeg/builds/ffmpeg-release-essentials.7z and place it (only the file ffmpeg.exe) in main.py directory to enable screen-recording." 
            print(message)
            self.log_system(message, job_id)
            return
            
        #written by AI
        
        os.makedirs(self.RECORDINGS_IN_PROGRESS_FOLDER, exist_ok=True)
        filename = f"{self.RECORDINGS_IN_PROGRESS_FOLDER}/{job_id}.mkv"

        drawtext = (
            f"drawtext=text='job_id  {job_id}':"
            "x=200:y=20:"
            "fontsize=32:"
            "fontcolor=lightyellow:"
            "box=1:"
            "boxcolor=black@0.5"
        )

        if platform.system() == "Windows":
            capture = ["-f", "gdigrab", "-i", "desktop"]
            ffmpeg = "./ffmpeg.exe"

            recording_process = subprocess.Popen(
                [ffmpeg, "-y", *capture, "-framerate", "15", "-vf", drawtext,
                "-vcodec", "libx264", "-preset", "ultrafast", filename],
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                creationflags=getattr(subprocess, "CREATE_NEW_PROCESS_GROUP", 0)
            )
        else:
            capture = ["-video_size", "1920x1080", "-f", "x11grab", "-i", ":0.0"]
            ffmpeg = "ffmpeg"
            recording_process = subprocess.Popen(
                [ffmpeg, "-y", *capture, "-framerate", "15", "-vf", drawtext,
                "-vcodec", "libx264", "-preset", "ultrafast", filename],
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                start_new_session=True
            )
        time.sleep(0.2) #adding dummy time to start the recording
        
        self.recording_process = recording_process  
        self.log_system("recording started", job_id)
  
    #stop recording
    def stop(self, job_id=None) -> None:
        # is allowed to do global KILL on all ffmpeg-processes
        #written by AI
        try:
            self.log_system("stop recording", job_id)
        except Exception: pass

        recording_process = self.recording_process
        self.recording_process = None

        try:
            if recording_process is not None:
                if platform.system() == "Windows":
                    try:
                        recording_process.send_signal(getattr(signal, "CTRL_BREAK_EVENT", signal.SIGTERM))
                    except Exception:
                        recording_process.terminate()

                    try:
                        recording_process.wait(timeout=8)
                    except subprocess.TimeoutExpired:
                        subprocess.run(
                            ["taskkill", "/IM", "ffmpeg.exe", "/T", "/F"],
                            stdout=subprocess.DEVNULL,
                            stderr=subprocess.DEVNULL,
                            check=False,
                        )

                else:
                    try:
                        os.killpg(recording_process.pid, signal.SIGINT)
                    except Exception:
                        recording_process.terminate()

                    try:
                        recording_process.wait(timeout=8)
                    except subprocess.TimeoutExpired:
                        subprocess.run(
                            ["killall", "-q", "-KILL", "ffmpeg"],
                            stdout=subprocess.DEVNULL,
                            stderr=subprocess.DEVNULL,
                            check=False,
                        )
            else:
                # fallback if proc-object tappats bort
                if platform.system() == "Windows":
                    subprocess.run(
                        ["taskkill", "/IM", "ffmpeg.exe", "/T", "/F"],
                        stdout=subprocess.DEVNULL,
                        stderr=subprocess.DEVNULL,
                        check=False,
                    )
                else:
                    subprocess.run(
                        ["killall", "-q", "-KILL", "ffmpeg"],
                        stdout=subprocess.DEVNULL,
                        stderr=subprocess.DEVNULL,
                        check=False,
                    )

        except Exception as err:
            print("WARN from stop():", err)

    #upload to a shared drive
    def upload_recording(self, job_id, max_attempts=3) -> bool:
    
        local_file = f"{self.RECORDINGS_IN_PROGRESS_FOLDER}/{job_id}.mkv"
        local_file = Path(local_file)

        remote_path = Path(self.RECORDINGS_DESTINATION_FOLDER) / f"{job_id}.mkv"
        remote_path.parent.mkdir(parents=True, exist_ok=True)

        for attempt in range(max_attempts):
            try:
                
                shutil.copy2(local_file, remote_path)
                #print(f"✓ Upload successful: {remote_path}")
                self.log_system(f"upload success: {remote_path}", job_id)
                try: os.remove(local_file)
                except Exception: pass

                return True

            except Exception as e:
                wait_time = (attempt + 1) ** 2
                print(f"Attempt {attempt+1}/{max_attempts} failed: {e}")
                time.sleep(wait_time)
        
        self.log_system(f"upload failed: {remote_path}", job_id)
        return False

    # cleanup aborted screen-recordings
    def cleanup_aborted_recordings(self):

        directory = Path(self.RECORDINGS_IN_PROGRESS_FOLDER)
        if not directory.exists():
            return
        
        for file in directory.iterdir():

            if file.is_file() and file.suffix == ".mkv":
                job_id = file.stem
                
                try:
                    self.upload_recording(job_id)
                    self.log_system(f"cleanup upload of {job_id}") #add recovery procedure for jo_id?
                except Exception as err:
                    self.log_system(f"cleanup failed for {job_id}: {err}")

# for access
class FriendsRepository:
    ''' example access-control source to use 'personal_inbox' '''

    def __init__(self, log_system) -> None:
        self.log_system = log_system
        self.access_by_email = {}
        self.access_file_mtime = None


    def ensure_friends_file_exists(self, path="friends.xlsx") -> None:
        ''' Makes a template if no friends.xlsx '''
        if os.path.exists(path):
            return

        wb = Workbook()
        ws = wb.active
        assert ws is not None #to satisfy pylance

        # headers
        ws["A1"] = "email"
        ws["B1"] = "ping"
        ws["C1"] = "job1"
        ws["D1"] = "job2"

        # rows
        ws["A2"] = "alice@example.com"
        ws["B2"] = "x"

        ws["A3"] = "bob@test.com"
        ws["B3"] = "x"
        ws["C3"] = "x"
        ws["D3"] = "x"

        wb.save(path)
        wb.close()
    

    def load_access_map(self, filepath="friends.xlsx") -> dict:
        #code written by AI
        '''
        Reads friends.xlsx and returns eg.:

        {
            "alice@example.com": {"ping"},
            "ex2@whatever.com": {"ping", "job1"}
        }

        Presumptions:
        A1 = email
        row 1 contains job_type
        'x' gives access
        '''
        wb = load_workbook(filepath, data_only=True)
        try:
            ws = wb.active
            assert ws is not None #to satisfy pylance

            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                raise ValueError("friends.xlsx contains no users")

            header = rows[0]
            access_map: dict[str, set[str]] = {}

            for row in rows[1:]:
                email_cell = row[0]
                if email_cell is None:
                    continue

                email = str(email_cell).strip().lower()
                if not email:
                    continue

                permissions = set()

                for col in range(1, len(header)):
                    jobname = header[col]
                    if jobname is None:
                        continue

                    jobname = str(jobname).strip().lower()
                    cell = row[col] if col < len(row) else None

                    if cell is None:
                        continue

                    if str(cell).strip().lower() == "x":
                        permissions.add(jobname)

                access_map[email] = permissions

            return access_map
        finally:
            wb.close()


    def reload_if_modified(self, force_reload=False, filepath="friends.xlsx") -> bool:
        #code written by AI
        '''      reload friends.xlsx if changed.       '''

        if not os.path.exists(filepath):
            raise FileNotFoundError(f"{filepath} not found")

        current_mtime = os.path.getmtime(filepath)

        if (not force_reload) and (self.access_file_mtime == current_mtime):
            return False   # ingen ändring

        new_access = self.load_access_map(filepath)

        self.access_by_email = new_access
        self.access_file_mtime = current_mtime

        return True


    def is_allowed_sender(self, email_address: str) -> bool:

        email_address = email_address.strip().lower()
        result = email_address in self.access_by_email
        self.log_system(f"returning: {result}")
        return result


    def has_job_access(self, email_address: str, job_type: str) -> bool:
        email_address = email_address.strip().lower()
        job_type = job_type.strip().lower()
        result = job_type in self.access_by_email.get(email_address, set())
        self.log_system(f"returning: {result}")
        return result

# for network check
class NetworkService:
    ''' checks if the computer is connected to company LAN '''

    #Example: r"G:\\" or r"\\\\server\\share"
    NETWORK_HEALTHCHECK_PATH = r"/"  #stub, this will always work


    def __init__(self, log_system) -> None:
        self.log_system = log_system
        self.network_state = False #assume offline at start
        self.next_network_check_time = 0


    def has_network_access(self) -> bool:
        #this runs at highest once every hour (if online), or before new jobs


        now = time.time()

        if now < self.next_network_check_time:
            return self.network_state

        try:
            os.listdir(self.NETWORK_HEALTHCHECK_PATH)
            online = True
            
        except Exception:
            online = False
            

        # update log if any network change (and UI? )
        if online != self.network_state:
            self.network_state = online

            if online:
                self.log_system("network restored")
            else:
                self.log_system(f"WARN: network lost")

        # check every minute if offline, else every hour (??)
        if online:
            self.next_network_check_time = now + 3600   # 1 h
        else:
            self.next_network_check_time = now + 60     # 1 min
        
        return online

# for audit-style log
class AuditRepository:
    ''' handles job_audit.db, an audit-style robot activity log '''
    def __init__(self, log_system) -> None:
        self.log_system = log_system
        

    def ensure_db_exists(self) -> None:
        
        with sqlite3.connect("job_audit.db") as conn:
            cur = conn.cursor()
           
            cur.execute('''
                CREATE TABLE IF NOT EXISTS audit_log
                         (
                        job_id INTEGER PRIMARY KEY, 
                        job_type TEXT, 
                        job_status TEXT, 
                        email_address TEXT, 
                        email_subject TEXT, 
                        source_ref TEXT,
                        job_start_date TEXT, 
                        job_start_time TEXT, 
                        job_finish_time TEXT, 
                        final_reply_sent INTEGER NOT NULL DEFAULT 0,
                        job_source_type TEXT,
                        error_code TEXT, 
                        error_message TEXT 
                        )
                        ''')
        conn.close()


    def insert_job(self, job_id, email_address=None, email_subject=None, source_ref=None, job_type: JobType | None=None, job_start_date=None, job_start_time=None, job_finish_time=None, job_status: JobStatus | None=None, final_reply_sent=None, job_source_type:JobSourceType | None=None, error_code=None, error_message=None,) -> None:
        # use for new row


        all_fields = {
            "job_id": job_id,
            "email_address": email_address,
            "email_subject": email_subject,
            "source_ref": source_ref,
            "job_type": job_type,
            "job_start_date": job_start_date,
            "job_start_time": job_start_time,
            "job_finish_time": job_finish_time,
            "job_status": job_status,
            "final_reply_sent": final_reply_sent,
            "job_source_type": job_source_type,
            "error_code": error_code,
            "error_message": error_message,
        }

        # ignore None:s
        fields = {k: v for k, v in all_fields.items() if v is not None}
        self.log_system(f"received fields: {fields}", job_id=job_id)
        
        columns = ", ".join(fields.keys())
        placeholders = ", ".join("?" for _ in fields)


        with sqlite3.connect("job_audit.db") as conn:
            cur = conn.cursor()

            cur.execute(
                f"INSERT INTO audit_log ({columns}) VALUES ({placeholders})",
                tuple(fields.values())
            )
        conn.close()


    def update_job(self, job_id, email_address=None, email_subject=None, source_ref=None, job_type: JobType | None=None, job_start_date=None, job_start_time=None, job_finish_time=None, job_status: JobStatus | None=None, final_reply_sent=None, job_source_type:JobSourceType | None=None, error_code=None, error_message=None,) -> None:
        # example use: self.audit_repo.update_job(job_id=20260311124501, job_type="job1")

        all_fields = {
            "job_id": job_id,
            "email_address": email_address,
            "email_subject": email_subject,
            "source_ref": source_ref,
            "job_type": job_type,
            "job_start_date": job_start_date,
            "job_start_time": job_start_time,
            "job_finish_time": job_finish_time,
            "job_status": job_status,
            "final_reply_sent": final_reply_sent,
            "job_source_type": job_source_type,
            "error_code": error_code,
            "error_message": error_message,
        }

        # ignore None-fields
        fields = {k: v for k, v in all_fields.items() if v is not None}
        self.log_system(f"received fields: {fields}", job_id=job_id)

        fields.pop("job_id", None)

        if not fields:
            return

        set_clause = ", ".join(f"{k}=?" for k in fields)


        with sqlite3.connect("job_audit.db") as conn:
            cur = conn.cursor()

            cur.execute(
                f"UPDATE audit_log SET {set_clause} WHERE job_id=?",
                (*fields.values(), job_id)
            )

            if cur.rowcount == 0:
                raise ValueError(f"update_job(): no row in DB with job_id={job_id}")
        conn.close()


    
    def count_done_jobs_today(self) -> int:
        # used for UI dash

        today = datetime.date.today().isoformat()

        with sqlite3.connect("job_audit.db") as conn:
            cur = conn.cursor()
            cur.execute('''
                SELECT COUNT(*)
                FROM audit_log
                WHERE job_start_date = ?
                AND job_status = 'DONE'
            ''', (today,))
            
            result = cur.fetchone()[0]
        conn.close()

        return result

    # used to send max one notification-response a day
    def has_sender_job_today(self, sender_mail, job_id) -> bool:    

        today = datetime.datetime.now().strftime("%Y-%m-%d")
        with sqlite3.connect("job_audit.db") as conn:
            cur = conn.cursor()

            cur.execute(
                '''
                SELECT COUNT(*)
                FROM audit_log
                WHERE job_start_date = ? AND email_address = ?
                ''',
                (today, sender_mail,)
            )

            jobs_today = cur.fetchone()[0]
        conn.close()

        self.log_system(f"returning: {jobs_today > 0}", job_id)

        return jobs_today > 0


    def has_been_processed_today(self, source_ref) -> bool:
        # used to avoid bad loops in schedule-jobs

        today = datetime.datetime.now().strftime("%Y-%m-%d")
        with sqlite3.connect("job_audit.db") as conn:
            cur = conn.cursor()

            cur.execute(
                '''
                SELECT COUNT(*)
                FROM audit_log
                WHERE job_start_date = ? AND source_ref = ?
                ''',
                (today, source_ref,)
            )

            jobs_today = cur.fetchone()[0]
        conn.close()

        #self.log_system(f"returning {source_ref} is  {jobs_today > 0}")
        return jobs_today > 0


    # used to avoid conflicting job_id
    def get_latest_job_id(self) -> int:
        with sqlite3.connect("job_audit.db") as conn:
            cur = conn.cursor()
            cur.execute('''
                SELECT job_id
                FROM audit_log
                ORDER BY job_id DESC
                LIMIT 1
            ''')
            row = cur.fetchone()
        conn.close()

        return row[0] if row is not None else 0


    def get_failed_jobs(self, days=7):
        # implement in UI dash ?
        with sqlite3.connect("job_audit.db") as conn:
            cur = conn.cursor()
            cur.execute('''
                SELECT job_id, email_address, job_type, error_code, error_message
                FROM audit_log
                WHERE job_status = 'FAIL'
                AND job_start_date >= date('now', '-' || ? || ' days')
                ORDER BY job_id DESC
            ''', (days,))
        res = cur.fetchall()
        conn.close()
        
        return res


    def get_pending_reply_jobs(self) -> list[dict]:
        job_source_type: JobSourceType = "personal_inbox" # typed

        with sqlite3.connect("job_audit.db") as conn:
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()
            cur.execute(
                '''
                SELECT job_id, email_address, email_subject, source_ref, job_status, error_code, error_message
                FROM audit_log
                WHERE job_source_type = ?
                AND COALESCE(final_reply_sent, 0) = 0
                ORDER BY job_id
                ''',
                (job_source_type,)
            )
            rows = cur.fetchall()
        conn.close()

        list_of_dicts = [dict(row) for row in rows]

        return list_of_dicts

# for crash-safe mode
class SafeStopController:
    def __init__(self, log_system, log_ui, recording_service, ui, mail_backend_personal, audit_repo, generate_job_id, friends_repo) -> None:
        self.log_system = log_system
        self.log_ui = log_ui
        self.recording_service = recording_service
        self.ui = ui
        self.mail_backend_personal = mail_backend_personal
        self.audit_repo = audit_repo
        self.generate_job_id = generate_job_id
        self.friends_repo = friends_repo
        self._safestop_entered = False


    def enter_safestop(self, reason) -> None:
        ''' safestop breaks the normal state-machine and handover.json logic '''
        
        if self._safestop_entered: return # re-entrancy protection
        self._safestop_entered = True 

        print("ROBOTRUNTIME CRASHED:\n", reason) 

        try: self.log_system(f"ROBOTRUNTIME CRASHED: {reason}")
        except Exception: pass

        try: self.recording_service.stop()
        except Exception: pass

        try: self.send_admin_alert(reason)
        except Exception: pass

        try: self.log_ui("CRASH! All automations halted. Admin is notified.", blank_line_before=True)
        except Exception: pass

        # follow policy to always reply on personal_inbox
        try:
            if self.audit_repo.get_pending_reply_jobs(): 
                try: self.recovery_answer()
                except Exception:
                    try: self.log_system("WARN: at least one user has a pending reply")
                    except Exception: print("WARN: at least one user has a pending reply")
        except Exception:
            try: self.log_system("WARN: unable to check if users has pending replies")
            except Exception: print("WARN: unable to check if users has pending replies")



        try: self.ui.tk_set_hide_recording_overlay()
        except Exception: pass

        try: self.ui.tk_set_status("safestop")
        except Exception: # rather kill than allow UI dash freeze
            try: self.ui.tk_set_shutdown()
            except: os._exit(1)
            
            time.sleep(3)
            os._exit(0)
        
        self.run_degraded_mode()


    def recovery_answer(self) -> None:
        jobs = self.audit_repo.get_pending_reply_jobs()

        for row in jobs:
            job_id = row["job_id"]
            source_ref = row["source_ref"]
            job_status = row["job_status"]
            error_message = row["error_message"]

            try:
                path = Path(source_ref)
                if not path.exists():
                    self.log_system(f"recovery skipped: missing processing file {source_ref}", job_id)
                    continue

                mail = self.mail_backend_personal.parse_mail_file(str(path))


                recovery_text = (
                    "\n\nThis is a recovery message because the robot was interrupted during processing.\n"
                    "The information above is based on the audit log.\n"
                    "This message is sent upon restart, regardless of when or whether the job was completed.\n"
                    "If the request is still needed, please resend it."
                )

                # check if screenrecodring available
                remote_path_prog = Path(RecordingService.RECORDINGS_IN_PROGRESS_FOLDER) / f"{job_id}.mkv"
                remote_path_dest = Path(RecordingService.RECORDINGS_DESTINATION_FOLDER) / f"{job_id}.mkv"
                
                for remote_path in (remote_path_prog, remote_path_dest):
                    if os.path.isfile(remote_path):
                        recovery_text += (
                            f"\nA screen recording is available for this job: >link stub:{remote_path}< "
                            "(it may be longer than usual due to the interruption).")
                        break

                if job_status == "DONE":
                    extra_subject = "DONE"
                    extra_body = f"Job completed successfully. Jobid {job_id}." + recovery_text
                elif job_status == "FAIL":
                    extra_subject = "FAIL"
                    extra_body = f"Job failed after processing. {error_message or ''}" + recovery_text
                elif job_status == "REJECTED":
                    extra_subject = "FAIL"
                    extra_body = f"{error_message or 'Your email could not be processed.'}" + recovery_text
                else:
                    self.log_system(f"recovery skipped: unexpected job_status={job_status}", job_id)
                    continue

                self.mail_backend_personal.reply_and_delete(
                    candidate=mail,
                    extra_subject=extra_subject,
                    extra_body=extra_body,
                    job_id=job_id,
                )

                self.audit_repo.update_job(
                    job_id=job_id,
                    final_reply_sent=True,
                )

                self.log_system(f"recovery reply sent", job_id)

            except Exception as err:
                self.log_system(f"recovery reply failed: {err}", job_id)


    def run_degraded_mode(self, restartflag="restart.flag") -> Never:
        ''' emergency mode to reply end-user emails in 'personal_inbox' and wait for restart. '''
        
        try: self.log_system("running")
        except Exception: pass
        
        while True:
            try:
                time.sleep(1)

                # check for restart flag
                if os.path.isfile(restartflag):
                    try: os.remove(restartflag)
                    except Exception: pass
                    try: self.log_system(f"restart-command received from {restartflag}")
                    except Exception: pass
                    self.restart_application()
                
                # reject all mail to 'personal_inbox', and check for system command
                paths = self.mail_backend_personal.fetch_from_inbox(max_items=1)
                if not paths:
                    continue
                
                inbox_path = paths[0]
                
                mail = self.mail_backend_personal.parse_mail_file(inbox_path)
                del inbox_path
                
                mail = self.mail_backend_personal.claim_to_processing(mail)
                self.log_ui(f"email from {mail.sender_email}", blank_line_before=True)

                if not self.friends_repo.is_allowed_sender(mail.sender_email):
                    self.log_ui("--> rejected (not in friends.xlsx)")
                    self.mail_backend_personal.delete_from_processing(mail)
                    continue
   
                if "restart1234" in mail.subject.strip().lower():
                    try: self.log_system(f"restart command received from {mail.sender_email}")
                    except Exception: pass
                    try: self.mail_backend_personal.reply_and_delete(mail, extra_subject="got it!", extra_body="Reboot command received")
                    except Exception: pass
                    self.restart_application()
                
                elif "stop1234" in mail.subject.strip().lower():
                    try: self.log_system(f"stop command received from {mail.sender_email}")
                    except Exception: pass
                    try: self.mail_backend_personal.reply_and_delete(mail, extra_subject="got it!", extra_body="Stop-command received, shutting down..")
                    except Exception: pass
                    try: self.ui.tk_set_shutdown()
                    except Exception: os._exit(1)
                    os._exit(0)

                
                try: self.mail_backend_personal.reply_and_delete(mail, extra_subject="FAIL", extra_body="Robot is out-of-service. Your email was deleted.")
                except Exception: pass
                try:
                    job_id = self.generate_job_id()
                    now = datetime.datetime.now()
                    job_source_type: JobSourceType = "personal_inbox" # typed 
                    
                    self.audit_repo.insert_job(
                        job_id=job_id,
                        email_address=mail.sender_email,
                        email_subject=mail.subject,
                        job_start_date=now.strftime("%Y-%m-%d"),
                        job_start_time=now.strftime("%H:%M:%S"),
                        job_status="REJECTED",
                        error_code="SAFESTOP",
                        job_source_type = job_source_type,
                    )
                except Exception: pass
                try: self.log_ui("--> rejected (safestop)")
                except Exception: pass
            
            except Exception as err:
                try: self.log_system(f"err: {err}")
                except Exception: pass


    def restart_application(self) -> Never:
        # this really works on UI dash freeze?
        try:
            self.ui.tk_set_shutdown()
        except Exception:
            pass

        try:
            subprocess.Popen(
                [sys.executable, *sys.argv],
                start_new_session=True,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                close_fds=True,
            )
        except Exception:
            os._exit(1)

        time.sleep(3)
        os._exit(0)


    def send_admin_alert(self, reason):
        mail=JobCandidate(
            source_ref="safestop, no real source_ref",
            sender_email="example_admin@company.com",
            subject="safestop entered",
            body=f"Reason: {reason}",
            job_source_type="personal_inbox", #?
            source_data={},
            )
         
        # maybe build a dedicated sender instead, since this is not a reply
        self.mail_backend_personal.send_reply(candidate=mail, extra_subject=mail.subject, extra_body=mail.body)


# ============================================================
# UI
# ============================================================

# for dashboard - "the face"
class DashboardUI:
    def __init__(self):
        bg_color ="#000000" #or "#111827"
        text_color = "#F5F5F5"

        self._build_root(bg_color)
        self._build_header(bg_color, text_color)
        self._build_body(bg_color, text_color)
        self._build_footer(bg_color, text_color)
        
        #self.debug_grid(self.root)


    def attach_runtime(self, robot_runtime) -> None:
        self.robot_runtime = robot_runtime


    def run(self) -> None:
        self.root.mainloop()


    def _build_root(self,bg_color):
        self.root = tk.Tk()
        self.root.geometry('1800x1000+0+0')
        #self.root.geometry('1800x200+0+0')
        #self.root.attributes("-fullscreen", True)
        self.root.resizable(False, False)

        self.root.configure(bg=bg_color, padx=50)
        self._closing = False
        self.root.protocol("WM_DELETE_WINDOW", self.shutdown)

        self.root.title('RPA dashboard')
        self._create_recording_overlay()

        # layout using grid
        self.root.grid_rowconfigure(1, weight=1)
        self.root.grid_columnconfigure(0, weight=1)


    def _build_header(self, bg_color, text_color):
        self.header = tk.Frame(self.root, bg=bg_color)
        
        self.header.grid(row=0, column=0, sticky="ew")
        self.header.grid_columnconfigure(2, weight=1)  
        self.header.grid_rowconfigure(0, weight=1)  

        # Header content
        self.rpa_text_label = tk.Label(self.header, text="RPA:", fg=text_color, bg=bg_color, font=("Arial", 100, "bold"))  #snyggare: "Segoe UI"
        self.rpa_text_label.grid(row=0, column=0, padx=16, pady=16, sticky="w")
        self.rpa_status_label = tk.Label(self.header, text="", fg="red", bg=bg_color, font=("Arial", 100, "bold"))
        self.rpa_status_label.grid(row=0, column=1, padx=16, pady=16, sticky="w")
        self.status_dot = tk.Label(self.header, text="", fg="#22C55E", bg=bg_color, font=("Arial", 50, "bold"))
        self.status_dot.grid(row=0, column=2, sticky="w")


        # jobs done today (counter + label in same grid)
        self.jobs_counter_frame = tk.Frame(self.header, bg=bg_color)
        self.jobs_counter_frame.grid(row=0, column=3, sticky="ne", padx=40, pady=30)
        self.jobs_counter_frame.grid_rowconfigure(0, weight=1)
        self.jobs_counter_frame.grid_columnconfigure(0, weight=1)


        # normal view (jobs done today)
        self.jobs_normal_view = tk.Frame(self.jobs_counter_frame, bg=bg_color)
        self.jobs_normal_view.grid(row=0, column=0, sticky="nsew")
        self.jobs_normal_view.grid_columnconfigure(0, weight=1)

        self.jobs_done_label = tk.Label(    self.jobs_normal_view,    text="0",    fg=text_color,    bg=bg_color,    font=("Segoe UI", 140, "bold"),       anchor="e",        justify="right")
        self.jobs_done_label.grid(row=0, column=0, sticky="e")

        self.jobs_counter_text = tk.Label(            self.jobs_normal_view,            text="jobs done today",            fg="#A0A0A0",            bg=bg_color,            font=("Arial", 14, "bold"),            anchor="e"        )
        self.jobs_counter_text.grid(row=1, column=0, sticky="e", pady=(0, 6))

        # safestop view (big X)
        self.jobs_error_view = tk.Frame(self.jobs_counter_frame, bg=bg_color)
        self.jobs_error_view.grid(row=0, column=0, sticky="nsew")

        self.safestop_x_label = tk.Label(            self.jobs_error_view,                        text="X",            bg="#DC2626",            fg="#FFFFFF",            font=("Segoe UI", 140, "bold")        ) #text="✖",
        self.safestop_x_label.pack(expand=True)


        # show normal view at startup
        self.jobs_normal_view.tkraise()

        # 'online'-status animation
        self._online_animation_after_id = None
        self._online_pulse_index = 0

        # 'working...'-status animation
        self._working_animation_after_id = None
        self._working_dots = 0


    def _build_body(self,bg_color, text_color):
        self.body = tk.Frame(self.root, bg=bg_color)        
        self.body.grid(row=1, column=0, sticky="nsew")
        self.body.grid_rowconfigure(0, weight=1)
        self.body.grid_columnconfigure(0, weight=1)

        # body content
        log_and_scroll_container = tk.Frame(self.body, bg=bg_color)
        log_and_scroll_container.grid(row=0, column=0, sticky="nsew")
        log_and_scroll_container.grid_rowconfigure(0, weight=1)
        log_and_scroll_container.grid_columnconfigure(0, weight=1)

        # the right-hand side scrollbar
        scrollbar = tk.Scrollbar(log_and_scroll_container, width=23, troughcolor="#0F172A", bg="#1E293B", activebackground="#475569", bd=0, highlightthickness=0, relief="flat")
        scrollbar.grid(row=0, column=1, sticky="ns")

        # the 'console'-style log
        self.log_text = tk.Text(log_and_scroll_container, yscrollcommand=scrollbar.set, bg=bg_color, fg=text_color, insertbackground="black", font=("DejaVu Sans Mono", 20), wrap="none", state="disabled", bd=0,highlightthickness=0) #glow highlightbackground="#1F2937", highlightthickness=1 
        self.log_text.grid(row=0, column=0, sticky="nsew")
        scrollbar.config(command=self.log_text.yview)


    def _build_footer(self,bg_color, text_color):
        self.footer = tk.Frame(self.root, bg=bg_color)        
        self.footer.grid(row=2, column=0, sticky="nsew")
        self.footer.grid_rowconfigure(0, weight=1)
        self.footer.grid_columnconfigure(0, weight=1)
        
        # footer content
        self.last_activity_label = tk.Label(self.footer, text="last activity: xx:xx", fg="#A0A0A0", bg=bg_color, font=("Arial", 14, "bold"), anchor="e")
        self.last_activity_label.grid(row=0, column=1, padx=8, pady=16)


    def debug_grid(self,widget):
        #highlights all grids with red
        for child in widget.winfo_children():
            try:
                child.configure(highlightbackground="red", highlightthickness=1)
            except Exception:
                pass
            self.debug_grid(child)


    def update_status_display(self, status: UIStatusText | None = None):
        # sets the status

        # stops any ongoing animations
        self._stop_online_animation()
        self._stop_working_animation()
        self.status_dot.config(text="")


        # changes text
        if status=="online":
            self.rpa_status_label.config(text="online", fg="#22C55E")
            self.jobs_normal_view.tkraise()
            self.status_dot.config(text="●")
            self._start_online_animation()
            
        elif status=="no network":
            self.rpa_status_label.config(text="no network", fg="red")
            self.jobs_normal_view.tkraise()
            
        elif status=="working":
            self.rpa_status_label.config(text="working...", fg="#FACC15")
            self.jobs_normal_view.tkraise()
            self._start_working_animation()

        elif status=="safestop":
            self.rpa_status_label.config(text="safestop", fg="red")
            self.jobs_error_view.tkraise()
            
        elif status=="ooo":
            self.rpa_status_label.config(text="out-of-office", fg="#FACC15")
            self.jobs_normal_view.tkraise()


    def set_jobs_done_today(self, n) -> None:
        self.jobs_done_label.config(text=str(n))


    def _create_recording_overlay(self) -> None:
        #written by AI
        self.recording_win = tk.Toplevel(self.root)
        self.recording_win.withdraw()                 # hidden at start
        self.recording_win.overrideredirect(True)    # no title/border
        self.recording_win.configure(bg="black")

        try: self.recording_win.attributes("-topmost", True)
        except Exception: pass

        width = 250
        height = 110
        x = self.root.winfo_screenwidth() - width - 30
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.recording_win.geometry(f"{width}x{height}+{x}+{y}")

        frame = tk.Frame(           self.recording_win,            bg="black",            highlightbackground="#444444",            highlightthickness=1,            bd=0        )
        frame.pack(fill="both", expand=True)

        canvas = tk.Canvas(        frame,        width=44,        height=44,        bg="black",        highlightthickness=0,        bd=0        )
        canvas.place(x=18, y=33)
        canvas.create_oval(4, 4, 40, 40, fill="#DC2626", outline="#DC2626")

        label = tk.Label(            frame,            text="RECORDING",            fg="#FFFFFF",            bg="black",            font=("Arial", 20, "bold"),            anchor="w"        )
        label.place(x=75, y=33)

        
    def show_recording_overlay(self) -> None:
        #written by AI
        try:
            width = 250
            height = 110
            x = self.root.winfo_screenwidth() - width - 30
            y = (self.root.winfo_screenheight() // 2) - (height // 2)
            self.recording_win.geometry(f"{width}x{height}+{x}+{y}")

            self.recording_win.deiconify()
            self.recording_win.lift()

            try:
                self.recording_win.attributes("-topmost", True)
            except Exception:
                pass
        except Exception:
            pass


    def hide_recording_overlay(self) -> None:
        # hides recording window
        try: self.recording_win.withdraw()
        except Exception: pass


    def _start_working_animation(self):
        if self._working_animation_after_id is None:
            self._animate_working()

    def _animate_working(self):
        #written by AI
        states = ["working", "working.", "working..", "working..."]
        self._working_dots = (self._working_dots + 1) % len(states)
        self.rpa_status_label.config(text=states[self._working_dots])
        self._working_animation_after_id = self.root.after(500, self._animate_working)

    def _stop_working_animation(self):
        if self._working_animation_after_id is not None:
            self.root.after_cancel(self._working_animation_after_id)
            self._working_animation_after_id = None
            self._working_dots = 0

    def _start_online_animation(self):
        if self._online_animation_after_id is None:
            self._online_pulse_index = 0
            self._animate_online()

    def _animate_online(self):
        # green puls animation
        colors = ["#22C55E", "#16A34A","#000000", "#15803D", "#16A34A"]
        color = colors[self._online_pulse_index]

        self.status_dot.config(fg=color)

        self._online_pulse_index = (self._online_pulse_index + 1) % len(colors)
        self._online_animation_after_id = self.root.after(1000, self._animate_online)

    def _stop_online_animation(self):
        if self._online_animation_after_id is not None:
            self.root.after_cancel(self._online_animation_after_id)
            self._online_animation_after_id = None

        
    def append_ui_log(self, log_line: str, blank_line_before: bool = False) -> None:
        # append the log

        self.log_text.config(state="normal") # open for edit
        now = datetime.datetime.now().strftime("%H:%M")

        if blank_line_before:
            self.log_text.insert("end", "\n")

        self.log_text.insert("end", f"[{now}] {log_line}\n")

        self.log_text.config(state="disabled") # closing edit
        self.log_text.see("end")


    def shutdown(self) -> Never | None:
        if self._closing: return
        self._closing = True

        try: self.robot_runtime.recording_service.stop()
        except Exception: pass

        self.root.destroy()

    # all 'tk_set_...' are wrappers
    def tk_set_status(self, status: UIStatusText) -> None:
        self.root.after(0, lambda: self.update_status_display(status))

    def tk_set_log(self, text: str, blank_line_before: bool = False) -> None:
        self.root.after(0, lambda: self.append_ui_log(text, blank_line_before))

    def tk_set_show_recording_overlay(self) -> None:
        self.root.after(0, self.show_recording_overlay)

    def tk_set_hide_recording_overlay(self) -> None:
        self.root.after(0, self.hide_recording_overlay)

    def tk_set_jobs_done_today(self, n: int) -> None:
        self.root.after(0, lambda: self.set_jobs_done_today(n))
    
    def tk_set_shutdown(self,) -> None:
        self.root.after(0, self.shutdown)


# ============================================================
# MAIN ENTRYPOINT
# ============================================================

# for core orchestration logic - "the brain"
class RobotRuntime:

    def __init__(self, ui):

        self.in_dev_mode = True

        self.ui = ui
        self.handover_repo = HandoverRepository(self.log_system)  #nu äger Runtime en handover_repo (som får med append-metod)
        self.friends_repo = FriendsRepository(self.log_system)
        self.audit_repo = AuditRepository(self.log_system)
        self.network_service = NetworkService(self.log_system)
        self.recording_service = RecordingService(self.log_system)
        self.safestop_controller = SafeStopController(self.log_system, self.log_ui, self.recording_service, ui, ExampleMailBackend(self.log_system, "personal_inbox"), self.audit_repo, self.generate_job_id, self.friends_repo) 
        self.job_handlers = {
            "ping": ExamplePingJobHandler(self.log_system),
            "job1": ExampleJob1Handler(self.log_system), 
            "job2": ExampleJob2Handler(self.log_system), 
            "job3": ExampleJob3Handler(self.log_system),}
    
        self.pre_handover_executor = PreHandoverExecutor(log_system=self.log_system, log_ui=self.log_ui, update_ui_status=self.update_ui_status, ui_dot_tk_set_show_recording_overlay=self.ui.tk_set_show_recording_overlay, generate_job_id=self.generate_job_id, recording_service=self.recording_service, audit_repo=self.audit_repo, in_dev_mode=self.in_dev_mode)
        self.scheduled_flow = ScheduledFlow(log_system=self.log_system, log_ui=self.log_ui, audit_repo=self.audit_repo, job_handlers=self.job_handlers, in_dev_mode=self.in_dev_mode, pre_handover_executor=self.pre_handover_executor)
        self.mail_flow = MailFlow(self.log_system, self.log_ui, self.friends_repo, self.is_within_operating_hours, self.network_service, self.job_handlers, self.pre_handover_executor)
        self.post_handover_finalizer = PostHandoverFinalizer(self.log_system, self.log_ui, self.audit_repo, self.job_handlers, self.recording_service, self.ui.tk_set_hide_recording_overlay, self.refresh_jobs_done_today_display, self.in_dev_mode)

        
    def initialize_runtime(self):
        
        VERSION = 0.4
        self.log_system(f"RuntimeThread started, version={VERSION}")

        self.handover_repo.write(ActiveJob(
            ipc_state="idle"
            )) # no-resume policy, always cold start

        # cleanup
        for fn in ["stop.flag", "restart.flag"]:
            try: os.remove(fn)
            except Exception: pass

        self.network_service.has_network_access()

        atexit.register(self.recording_service.stop) #extra protection during normal python exit
        self.recording_service.stop() #stop any remaing recordings
        self.recording_service.cleanup_aborted_recordings()

        self.friends_repo.ensure_friends_file_exists()

        self.friends_repo.reload_if_modified(force_reload=True)

        self.audit_repo.ensure_db_exists()

        self.refresh_jobs_done_today_display()

        # for unanswered personal_inbox email
        if self.audit_repo.get_pending_reply_jobs():
            self.safestop_controller.recovery_answer()


    def run(self) -> None:
        self.initialize_runtime()

        self.prev_ui_status = None
        prev_ipc_state = None
        watchdog_started_at = None
        watchdog_timeout = 600 #600 for 10 min
        if self.in_dev_mode: watchdog_timeout = 10

        poll_interval = 1   # inverval for each cycle    

        while True:
            try:
                
                handover_data = self.handover_repo.read()
                ipc_state = handover_data.get("ipc_state")
                
                #dispatch
                if ipc_state == "idle":             # Orchestrator owns workflow on "idle"
                    self.check_for_jobs()
                    time.sleep(poll_interval)

                elif ipc_state == "job_queued":     # signal to RPA tool (from Orchestrator) to take workflow
                    time.sleep(poll_interval)

                elif ipc_state == "job_running":    # signal from RPA tool that workflow is taken
                    time.sleep(poll_interval)

                elif ipc_state == "job_verifying":  # signal to orchestrator (from RPA tool) to re-take workflow
                    self.finalize_current_job(handover_data)

                elif ipc_state == "safestop":       # signal to orchestrator (from RPA tool) to crash due to an error 
                    raise RuntimeError(f"crash signal received from RPA tool for job_id: {handover_data.get('job_id')}")
                    

                # log all ipc_state transitions
                if ipc_state != prev_ipc_state:
                    self.update_ui_status(ipc_state)
                    self.log_system(f"state transition detected by CPU-poll: {prev_ipc_state} -> {ipc_state}")
                    print(f"state transition detected by CPU-poll: {prev_ipc_state} -> {ipc_state}")

                    # note handover time or last RPA tool activity
                    if ipc_state in ("job_queued", "job_running"):
                        watchdog_started_at = time.time()
                    else:
                        watchdog_started_at = None
                   
                    # update DB when/if RPA tool starts the job
                    if ipc_state == "job_running":
                        self.audit_repo.update_job(job_id=handover_data.get("job_id"), job_status="RUNNING")
                    

                # initiate crash if RPA tool takes too long (to start or finish)
                if watchdog_started_at and ipc_state in ("job_queued", "job_running") and time.time() - watchdog_started_at > watchdog_timeout:
                    self.audit_repo.update_job(
                        job_id=handover_data.get("job_id"),
                        job_status="FAIL",
                        error_code="TIMEOUT",
                        error_message="No progress for 10 minutes",
                    )
                    watchdog_started_at = None
                    raise RuntimeError(f"job_id {handover_data.get('job_id')} crashed: no lifesign from RPA tool for {watchdog_timeout/60} minutes") #safe-stop policy
                
                prev_ipc_state = ipc_state


            except Exception:  # safe-stop policy on crash
                err = traceback.format_exc()
                #self.handover_repo.write(ActiveJob( ipc_state="safestop") )
                self.safestop_controller.enter_safestop(reason=err) 



    def refresh_jobs_done_today_display(self):
        # in UI dash

        count = self.audit_repo.count_done_jobs_today()
        self.ui.tk_set_jobs_done_today(count)

    def update_ui_status(self, ipc_state=None, forced_status=None) -> None:
               
        if forced_status is not None:
            if forced_status not in get_args(UIStatusText):
                raise ValueError(f"unknown forced_status: {forced_status}")
            ui_status: UIStatusText = forced_status

        else:
            if ipc_state is not None and ipc_state not in get_args(IpcState):
                raise ValueError(f"unknown ipc_state: {ipc_state}")

            if ipc_state == "safestop":
                ui_status = "safestop"

            elif ipc_state in ("job_queued", "job_running", "job_verifying"):
                ui_status = "working"

            elif self.network_service.network_state is False:
                ui_status = "no network"

            elif not self.is_within_operating_hours():
                ui_status = "ooo"

            else:
                ui_status = "online"

        if self.prev_ui_status != ui_status:
            self.ui.tk_set_status(ui_status)
            self.prev_ui_status = ui_status


    def log_ui(self, text:str, blank_line_before: bool = False) -> None:
        
        self.ui.tk_set_log(text, blank_line_before)
        

    def log_system(self, event_text: str, job_id=None, file="system.log"):

        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # get caller function name
        try:
            frame = sys._getframe(1)
            caller_name = frame.f_code.co_name
            instance = frame.f_locals.get("self")
            if instance is not None:
                class_name = instance.__class__.__name__
                caller = f"{class_name}.{caller_name}()"
            else:
                caller = f"{caller_name}()"

        except Exception:
            caller = "unknown_caller()"

        job_part = f" | JOB {job_id}" if job_id else ""
        log_line = f"{timestamp}{job_part} | {caller} | {event_text}"

        # normalize to single-line log
        log_line = " ".join(str(log_line).split())

        last_err = None
        for i in range(7):
            try:
                with open(file, "a", encoding="utf-8") as f:
                    f.write(log_line + "\n")
                    f.flush()
                return

            except Exception as err:
                last_err = err
                print(f"WARN: retry {i+1}/7 from log_system():", err)
                time.sleep(i + 1)

        # policy: allow system to work without log?
        print(f"WARN: failed after 7 attempts: {last_err}")  
 

    def check_for_jobs(self) -> bool:
        
        # 1. Mail first (priority)
        mail_result = self.mail_flow.poll_once()
        if mail_result.handover_data is not None:
            self.handover_repo.write(mail_result.handover_data)
            return True
        
        if mail_result.handled_anything:  # allow mail to starve scheduled
            return True   

        # 2. Scheduled jobs
        scheduled_result = self.scheduled_flow.poll_once()
        if scheduled_result.handover_data is not None:
            self.handover_repo.write(scheduled_result.handover_data)
            return True
   

        return False


    def generate_job_id(self) -> int:
        ''' unique id for all jobs '''

        job_id = int(datetime.datetime.now().strftime("%Y%m%d%H%M%S"))
        
        # simple busy-wait
        while self.audit_repo.get_latest_job_id() >= job_id:
            time.sleep(1)
            job_id = int(datetime.datetime.now().strftime("%Y%m%d%H%M%S"))

        self.log_system(f"assigned job_id", job_id)
        return job_id

    
    def is_within_operating_hours(self) -> bool:
        
        now = datetime.datetime.now().time()
        result = datetime.time(5,0) <= now <= datetime.time(23,0) # eg. working hours 05:00 to 23:00
        
        self.log_system(f"returning: {result}")
        return result
        

    def finalize_current_job(self, handover_data) -> None:
        
        self.post_handover_finalizer.poll_once(handover_data)

        self.handover_repo.write(ActiveJob(
            ipc_state="idle",
        ))


    def poll_for_stop_flag(self, stopflag="stop.flag"):
        # to stop python on operator manual stop on RPA tool

        self.log_system("poll_for_stop_flag() alive")

        while True:
            time.sleep(1)
            
            if os.path.isfile(stopflag):
                try: os.remove(stopflag)
                except Exception: pass

                try: self.log_system(f"found {stopflag}")
                except Exception: pass
                
                try: self.ui.tk_set_shutdown() #request soft-exit
                except Exception: os._exit(1)
                
                time.sleep(1)
                os._exit(0)  #kill if still alive after 1 sec 

def main() -> None:
    #run dashboard in main thred and 'the rest' in async worker
    ui = DashboardUI()
    robot_runtime = RobotRuntime(ui)
    ui.attach_runtime(robot_runtime)

    threading.Thread(target=robot_runtime.run, daemon=True).start() # 'the rest'
    threading.Thread(target=robot_runtime.poll_for_stop_flag, daemon=True).start() # killswitch triggered by RPA tool stop

    ui.run()


if __name__ == "__main__":
    main()
